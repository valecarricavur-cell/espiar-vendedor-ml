"""
espiar_vendedor.py
------------------
Monitorea las publicaciones activas de un vendedor de MercadoLibre Argentina.
Usa Playwright para recorrer las páginas de la tienda y requests para obtener
el detalle de cada publicación (vendidos, stock, marca, tipo de producto).

Genera dos salidas:
  - Excel (.xlsx): hoja principal + resumen por marca + detalle de variantes
  - Página HTML  : dashboard visual que se abre en el navegador

Uso:
    python espiar_vendedor.py <NICKNAME> [--carpeta NOMBRE]

    Ejemplo:
        python espiar_vendedor.py todoairelibregd --carpeta TODOAIRELIBRE

Dependencias:
    pip install playwright requests openpyxl python-dotenv
    playwright install chromium

Variables de entorno opcionales (archivo .env) para WhatsApp:
    TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, TWILIO_WHATSAPP_FROM, TWILIO_WHATSAPP_TO
"""

import os
import sys
import re
import glob
import time
import argparse
import textwrap
import requests
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    from twilio.rest import Client as TwilioClient
    TWILIO_DISPONIBLE = True
except ImportError:
    TWILIO_DISPONIBLE = False

# ─── Configuración ────────────────────────────────────────────────────────────

TWILIO_SID   = os.getenv("TWILIO_ACCOUNT_SID")
TWILIO_TOKEN = os.getenv("TWILIO_AUTH_TOKEN")
WA_FROM      = os.getenv("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886")
WA_TO        = os.getenv("TWILIO_WHATSAPP_TO")

REPORTS_ROOT   = Path("reportes_ml")
WORKERS        = 15   # requests paralelos para páginas de producto
DELAY_LISTING  = 1.5  # segundos entre páginas del listado (Playwright)

_HEADERS_WEB = {
    "User-Agent":      "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Accept-Language": "es-AR,es;q=0.9,en;q=0.8",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
}


# ─── Scraping: páginas del listado (Playwright) ───────────────────────────────

def _wid_desde_url(url: str) -> Optional[str]:
    """Extrae el wid (item ID individual del vendedor) de la URL del poly-card."""
    m = re.search(r'[?&]wid=(MLA\d+)', url)
    return m.group(1) if m else None


def _precio_desde_texto(texto: str) -> float:
    """Convierte '52.673' o '52673' a 52673.0"""
    limpio = re.sub(r'[^\d,]', '', texto).replace(',', '.')
    try:
        return float(limpio.replace('.', '').replace(',', '.')) if limpio else 0.0
    except ValueError:
        return 0.0


def scrape_paginas_vendedor(nickname: str) -> list[dict]:
    """
    Usa Playwright (headless) para recorrer TODAS las páginas de la tienda.
    Intenta primero con headless=True; si no encuentra productos, abre el
    navegador visible (headless=False) y reintenta.

    Retorna lista de dicts con: titulo, precio, item_id, url_catalogo,
    url_articulo, envio_gratis, marca_card (del card, no del spec).
    """
    from playwright.sync_api import sync_playwright

    base_url = f"https://listado.mercadolibre.com.ar/pagina/{nickname}/"

    def _extraer_cards(page) -> list[dict]:
        cards = page.query_selector_all(".poly-card")
        resultados = []
        for card in cards:
            try:
                link_el = card.query_selector("a.poly-component__title, a[href*='mercadolibre']")
                href = link_el.get_attribute("href") if link_el else ""

                title_el = card.query_selector(".poly-component__title")
                title = (title_el.inner_text() or "").strip()

                # Precio: buscar el monto actual (no el tachado)
                price_el = card.query_selector(
                    ".poly-price__current .andes-money-amount__fraction"
                )
                if not price_el:
                    price_el = card.query_selector(".andes-money-amount__fraction")
                price_text = price_el.inner_text() if price_el else "0"
                precio = _precio_desde_texto(price_text)

                shipping_el = card.query_selector(".poly-component__shipping")
                envio = "gratis" in (shipping_el.inner_text() if shipping_el else "").lower()

                # Marca del card (texto del seller badge, e.g. "MONTAGNE")
                seller_el = card.query_selector(".poly-component__seller")
                marca_card = (seller_el.inner_text() or "").strip() if seller_el else ""

                # Vendidos mostrados en el card (e.g. "+100 vendidos")
                sold_el = card.query_selector(".poly-component__sold")
                sold_text = (sold_el.inner_text() or "").strip() if sold_el else ""

                wid = _wid_desde_url(href)
                url_articulo = (
                    f"https://articulo.mercadolibre.com.ar/{wid[:3]}-{wid[3:]}"
                    if wid else ""
                )

                resultados.append({
                    "titulo":       title,
                    "precio":       precio,
                    "item_id":      wid or "",
                    "url_catalogo": href.split("?")[0] if href else "",
                    "url_articulo": url_articulo,
                    "envio_gratis": envio,
                    "marca_card":   marca_card,
                    "sold_card":    sold_text,
                })
            except Exception:
                continue
        return resultados

    def _scrape_con_headless(headless: bool) -> list[dict]:
        items = []
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless, args=["--no-sandbox"])
            ctx = browser.new_context(
                user_agent=_HEADERS_WEB["User-Agent"],
                locale="es-AR",
                extra_http_headers={"Accept-Language": "es-AR,es;q=0.9"},
            )
            page = ctx.new_page()

            offset = 0
            total_esperado = None

            while True:
                if offset == 0:
                    url = base_url
                else:
                    url = f"{base_url}_Desde_{offset + 1}_NoIndex_True"

                page.goto(url, wait_until="domcontentloaded", timeout=45000)
                page.wait_for_timeout(2500 if headless else 3500)

                # Leer total de resultados en la primera página
                if total_esperado is None:
                    total_el = page.query_selector(".ui-search-search-result__quantity-results")
                    if total_el:
                        t = re.search(r'([\d.]+)', total_el.inner_text().replace(".", ""))
                        total_esperado = int(t.group(1)) if t else 9999
                    else:
                        total_esperado = 9999

                cards = _extraer_cards(page)
                if not cards:
                    break

                items.extend(cards)
                print(f"      Página {offset//48 + 1}: {len(cards)} productos ({len(items)}/{total_esperado})")

                offset += 48
                if len(items) >= total_esperado:
                    break

                time.sleep(DELAY_LISTING)

            browser.close()
        return items

    print("      Intentando con navegador invisible…")
    items = _scrape_con_headless(headless=True)

    if not items:
        print("      Sin resultados headless — abriendo navegador visible…")
        items = _scrape_con_headless(headless=False)

    return items


# ─── Scraping: detalle de cada publicación (requests) ────────────────────────

def _limpiar_tipo(domain_id: str) -> str:
    return domain_id.replace("MLA-", "").replace("_", " ").title() if domain_id else "—"


def scrape_item_detalle(item: dict) -> dict:
    """
    Descarga la página individual del ítem y extrae vendidos, stock, marca y tipo.
    Modifica el dict in-place y lo retorna.
    """
    url = item.get("url_articulo", "")
    if not url:
        item.setdefault("vendidos", 0)
        item.setdefault("stock", None)
        item.setdefault("marca", item.get("marca_card", "Sin marca") or "Sin marca")
        item.setdefault("tipo", "—")
        item.setdefault("condicion", "new")
        item.setdefault("dias_activo", 1)
        item.setdefault("ventas_hist_dia", 0.0)
        item.setdefault("ventas_rec_dia", None)
        item.setdefault("variantes", [])
        item.setdefault("permalink", item.get("url_catalogo", ""))
        return item

    try:
        r = requests.get(url, headers=_HEADERS_WEB, timeout=20, allow_redirects=True)
        html = r.text

        # Vendidos
        m = re.search(r'(\d[\d.]*)\s+vendidos?', html, re.IGNORECASE)
        vendidos = int(m.group(1).replace(".", "")) if m else 0

        # Stock disponible
        m = re.search(r'available_quantity[\":\s]+(\d+)', html)
        stock = int(m.group(1)) if m else None

        # Marca desde tabla de specs
        m = re.search(
            r'>Marca<.*?class="andes-table__column--value"[^>]*>([^<]+)</span>',
            html, re.DOTALL
        )
        marca = (m.group(1).strip() if m else item.get("marca_card", "")) or "Sin marca"

        # Tipo de producto desde domain_id
        domain_ids = re.findall(r'"domain_id":\s*"([^"]+)"', html)
        tipo = _limpiar_tipo(domain_ids[0]) if domain_ids else "—"

        # Condición
        cond_hits = re.findall(r'"condition":\s*"([^"]+)"', html)
        condicion = cond_hits[0] if cond_hits else "new"

    except Exception as e:
        vendidos, stock, marca, tipo, condicion = 0, None, item.get("marca_card", "Sin marca"), "—", "new"

    # Actualizar el dict
    item["vendidos"]       = vendidos
    item["stock"]          = stock
    item["marca"]          = marca
    item["tipo"]           = tipo
    item["condicion"]      = condicion
    item["dias_activo"]    = 1         # sin fecha de creación disponible
    item["ventas_hist_dia"] = vendidos  # usamos total vendidos como proxy del historial
    item["ventas_rec_dia"]  = None      # se calcula al comparar snapshots
    item["variantes"]      = []
    item["permalink"]      = item.get("url_articulo") or item.get("url_catalogo", "")

    return item


def enriquecer_items_paralelo(raw_items: list[dict]) -> list[dict]:
    """Llama scrape_item_detalle en paralelo para todos los ítems."""
    total = len(raw_items)
    resultados = [None] * total
    completados = [0]

    with ThreadPoolExecutor(max_workers=WORKERS) as executor:
        futuros = {executor.submit(scrape_item_detalle, item): i
                   for i, item in enumerate(raw_items)}
        for fut in as_completed(futuros):
            idx = futuros[fut]
            try:
                resultados[idx] = fut.result()
            except Exception:
                resultados[idx] = raw_items[idx]
            completados[0] += 1
            if completados[0] % 50 == 0:
                print(f"      {completados[0]}/{total} ítems procesados…")

    return [r for r in resultados if r is not None]


# ─── Snapshot anterior ─────────────────────────────────────────────────────────

def cargar_reporte_anterior(nickname: str, carpeta: Path) -> tuple:
    """
    Lee el penúltimo Excel del vendedor (si existe).
    Retorna ({item_id: {precio, stock, vendidos}}, fecha_snapshot_anterior).
    """
    archivos = sorted(glob.glob(str(carpeta / f"{nickname}_*.xlsx")))
    if len(archivos) < 2:
        return {}, None

    ruta = archivos[-2]
    print(f"[i] Comparando contra: {Path(ruta).name}")

    stem   = Path(ruta).stem
    partes = stem.split("_")
    try:
        fecha_ant = datetime.strptime(f"{partes[-2]}_{partes[-1]}", "%Y%m%d_%H%M%S")
    except (ValueError, IndexError):
        fecha_ant = None

    wb = openpyxl.load_workbook(ruta)
    ws = wb.active
    encabezados = {cell.value: cell.column for cell in ws[1]}
    col_id  = encabezados.get("ID", 1)
    col_pre = encabezados.get("Precio (ARS)", 5)
    col_stk = encabezados.get("Stock Total", 6)
    col_ven = encabezados.get("Vendidos Totales", 7)

    datos = {}
    for fila in ws.iter_rows(min_row=2, values_only=True):
        iid = fila[col_id - 1]
        if iid:
            datos[str(iid)] = {
                "precio":   fila[col_pre - 1],
                "stock":    fila[col_stk - 1],
                "vendidos": fila[col_ven - 1],
            }
    return datos, fecha_ant


# ─── Comparación de snapshots ──────────────────────────────────────────────────

def comparar_snapshots(
    actuales:     list[dict],
    anteriores:   dict,
    fecha_actual: datetime,
    fecha_ant:    Optional[datetime],
) -> dict:
    if not anteriores:
        return {
            "cambios_precio":       [],
            "cambios_stock":        [],
            "ventas_estimadas":     0,
            "ventas_rec_dia_total": None,
            "delta_dias":           None,
            "nuevos":               [i["item_id"] for i in actuales],
            "eliminados":           [],
        }

    delta_dias = (
        max(0.01, (fecha_actual - fecha_ant).total_seconds() / 86400)
        if fecha_ant else None
    )

    ids_actuales   = {i["item_id"] for i in actuales}
    ids_anteriores = set(anteriores.keys())
    cambios_precio, cambios_stock = [], []
    ventas_est = 0

    for item in actuales:
        iid  = item["item_id"]
        prev = anteriores.get(iid)
        if not prev:
            continue

        p_ant, p_act = (prev["precio"] or 0), (item["precio"] or 0)
        if p_ant and p_act and p_ant != p_act:
            pct = (p_act - p_ant) / p_ant * 100
            cambios_precio.append({
                "id": iid, "titulo": item["titulo"],
                "antes": p_ant, "ahora": p_act, "delta%": round(pct, 1),
            })

        s_ant, s_act = (prev["stock"] or 0), (item["stock"] or 0)
        if s_ant != s_act:
            cambios_stock.append({
                "id": iid, "titulo": item["titulo"],
                "antes": s_ant, "ahora": s_act,
            })

        delta_v = max(0, (item["vendidos"] or 0) - (prev["vendidos"] or 0))
        ventas_est += delta_v
        if delta_dias:
            item["ventas_rec_dia"] = round(delta_v / delta_dias, 2)

    ventas_rec_dia_total = round(ventas_est / delta_dias, 2) if delta_dias else None

    return {
        "cambios_precio":        cambios_precio,
        "cambios_stock":         cambios_stock,
        "ventas_estimadas":      ventas_est,
        "ventas_rec_dia_total":  ventas_rec_dia_total,
        "delta_dias":            round(delta_dias, 1) if delta_dias else None,
        "nuevos":                list(ids_actuales - ids_anteriores),
        "eliminados":            list(ids_anteriores - ids_actuales),
    }


# ─── Excel ─────────────────────────────────────────────────────────────────────

AZUL   = "1A73E8"
VERDE  = "0F9D58"
GRIS   = "F5F5F5"
BLANCO = "FFFFFF"


def _estilo_encabezado(ws, fila: int, cols: list, color: str = AZUL):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color=BLANCO)
    for col, texto in enumerate(cols, start=1):
        c = ws.cell(row=fila, column=col, value=texto)
        c.fill, c.font = fill, font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 20


def _autowidth(ws, max_w=70):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, max_w)


def guardar_excel(
    items: list[dict],
    nickname: str,
    fecha: datetime,
    carpeta: Path,
    cambios: dict,
) -> Path:
    ts   = fecha.strftime("%Y%m%d_%H%M%S")
    ruta = carpeta / f"{nickname}_{ts}.xlsx"
    wb   = openpyxl.Workbook()

    fill_alt = PatternFill("solid", fgColor=GRIS)

    # ── Hoja 1: Publicaciones ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Publicaciones"
    cols1 = [
        "ID", "Título", "Marca", "Tipo de Producto",
        "Precio (ARS)", "Stock Total", "Vendidos Totales",
        "Ventas Rec/Día", "Condición", "Envío Gratis", "URL",
    ]
    _estilo_encabezado(ws1, 1, cols1)

    for r, it in enumerate(items, start=2):
        stock_str = it["stock"] if it["stock"] is not None else "N/D"
        fila = [
            it["item_id"], it["titulo"], it["marca"], it["tipo"],
            it["precio"], stock_str, it["vendidos"],
            it["ventas_rec_dia"] if it["ventas_rec_dia"] is not None else "—",
            it["condicion"], "Sí" if it["envio_gratis"] else "No", it["permalink"],
        ]
        for c, val in enumerate(fila, start=1):
            celda = ws1.cell(row=r, column=c, value=val)
            if r % 2 == 0:
                celda.fill = fill_alt

    ws1.freeze_panes = "A2"
    _autowidth(ws1)

    # ── Hoja 2: Por Marca ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Por Marca")
    cols2 = ["Marca", "Publicaciones", "Stock Total", "Vendidos Totales", "Ventas Rec/Día"]
    _estilo_encabezado(ws2, 1, cols2, VERDE)

    marcas: dict = defaultdict(lambda: {"pubs": 0, "stock": 0, "vendidos": 0, "rec": 0.0, "rc": 0})
    for it in items:
        m = it["marca"]
        marcas[m]["pubs"]     += 1
        marcas[m]["stock"]    += (it["stock"] or 0)
        marcas[m]["vendidos"] += it["vendidos"]
        if it["ventas_rec_dia"] is not None:
            marcas[m]["rec"] += it["ventas_rec_dia"]
            marcas[m]["rc"]  += 1

    for r, (marca, d) in enumerate(
        sorted(marcas.items(), key=lambda x: x[1]["vendidos"], reverse=True), start=2
    ):
        rec = round(d["rec"], 2) if d["rc"] else "—"
        for c, val in enumerate([marca, d["pubs"], d["stock"], d["vendidos"], rec], start=1):
            celda = ws2.cell(row=r, column=c, value=val)
            if r % 2 == 0:
                celda.fill = fill_alt

    ws2.freeze_panes = "A2"
    _autowidth(ws2)

    # ── Hoja 3: Por Tipo de Producto ───────────────────────────────────────────
    ws3 = wb.create_sheet("Por Tipo")
    cols3 = ["Tipo de Producto", "Publicaciones", "Stock Total", "Vendidos Totales"]
    _estilo_encabezado(ws3, 1, cols3, "7B1FA2")

    tipos: dict = defaultdict(lambda: {"pubs": 0, "stock": 0, "vendidos": 0})
    for it in items:
        t = it["tipo"]
        tipos[t]["pubs"]     += 1
        tipos[t]["stock"]    += (it["stock"] or 0)
        tipos[t]["vendidos"] += it["vendidos"]

    for r, (tipo, d) in enumerate(
        sorted(tipos.items(), key=lambda x: x[1]["vendidos"], reverse=True), start=2
    ):
        for c, val in enumerate([tipo, d["pubs"], d["stock"], d["vendidos"]], start=1):
            celda = ws3.cell(row=r, column=c, value=val)
            if r % 2 == 0:
                celda.fill = fill_alt

    ws3.freeze_panes = "A2"
    _autowidth(ws3)

    wb.save(ruta)
    print(f"[✓] Excel guardado: {ruta}")
    return ruta


# ─── HTML ──────────────────────────────────────────────────────────────────────

def generar_html(
    items: list[dict],
    nickname: str,
    fecha: datetime,
    cambios: dict,
    carpeta: Path,
) -> Path:
    ts   = fecha.strftime("%Y%m%d_%H%M%S")
    ruta = carpeta / f"{nickname}_{ts}.html"

    total_stock    = sum((i["stock"] or 0) for i in items)
    total_vendidos = sum(i["vendidos"] for i in items)
    precio_prom    = (
        sum(i["precio"] for i in items if i["precio"]) / len(items) if items else 0
    )

    marcas: dict = defaultdict(lambda: {"pubs": 0, "vendidos": 0, "rec": 0.0, "rc": 0})
    for it in items:
        m = it["marca"]
        marcas[m]["pubs"]     += 1
        marcas[m]["vendidos"] += it["vendidos"]
        if it["ventas_rec_dia"] is not None:
            marcas[m]["rec"] += it["ventas_rec_dia"]
            marcas[m]["rc"]  += 1

    top_marcas = sorted(marcas.items(), key=lambda x: x[1]["vendidos"], reverse=True)[:12]

    filas_marcas = ""
    for i, (m, d) in enumerate(top_marcas, 1):
        rec = f"{d['rec']:.1f}" if d["rc"] else "—"
        filas_marcas += (
            f"<tr><td>{i}</td><td><strong>{m}</strong></td>"
            f"<td>{d['pubs']}</td><td>{d['vendidos']:,}</td><td>{rec}</td></tr>\n"
        )

    top_items = sorted(items, key=lambda x: x["vendidos"], reverse=True)[:300]
    filas_items = ""
    for it in top_items:
        rec   = f"{it['ventas_rec_dia']:.2f}" if it["ventas_rec_dia"] is not None else "—"
        env   = "✅" if it["envio_gratis"] else "❌"
        stock = it["stock"] if it["stock"] is not None else "N/D"
        filas_items += (
            f"<tr>"
            f"<td><a href='{it['permalink']}' target='_blank'>{it['titulo'][:65]}</a></td>"
            f"<td>{it['marca']}</td>"
            f"<td>{it['tipo']}</td>"
            f"<td>${it['precio']:,.0f}</td>"
            f"<td>{stock}</td>"
            f"<td>{it['vendidos']:,}</td>"
            f"<td>{rec}</td>"
            f"<td>{env}</td>"
            f"</tr>\n"
        )

    ventas_est_str = (
        f"{cambios['ventas_estimadas']:,} uds en {cambios['delta_dias']} días "
        f"(≈ <strong>{cambios['ventas_rec_dia_total']}</strong>/día)"
        if cambios.get("ventas_rec_dia_total") is not None
        else "Primer snapshot — sin comparación aún"
    )

    html = textwrap.dedent(f"""<!DOCTYPE html>
    <html lang="es">
    <head>
      <meta charset="UTF-8">
      <title>Reporte ML — {nickname}</title>
      <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: 'Segoe UI', sans-serif; background: #f0f2f5; color: #333; }}
        header {{ background: #1A73E8; color: white; padding: 20px 32px; }}
        header h1 {{ font-size: 1.4rem; }}
        header p  {{ font-size: .85rem; opacity: .85; margin-top: 4px; }}
        .cards {{ display: flex; flex-wrap: wrap; gap: 16px; padding: 24px 32px 0; }}
        .card {{ background: white; border-radius: 10px; padding: 16px 24px;
                 box-shadow: 0 1px 4px rgba(0,0,0,.1); flex: 1; min-width: 160px; }}
        .card .val {{ font-size: 1.8rem; font-weight: 700; color: #1A73E8; }}
        .card .lbl {{ font-size: .78rem; color: #888; margin-top: 4px; }}
        section {{ margin: 24px 32px; background: white; border-radius: 10px;
                   box-shadow: 0 1px 4px rgba(0,0,0,.1); overflow: hidden; }}
        section h2 {{ background: #f8f9fa; padding: 14px 20px; font-size: 1rem;
                      border-bottom: 1px solid #e0e0e0; }}
        table {{ width: 100%; border-collapse: collapse; font-size: .82rem; }}
        th {{ background: #1A73E8; color: white; padding: 9px 12px; text-align: left; }}
        td {{ padding: 8px 12px; border-bottom: 1px solid #f0f0f0; }}
        tr:hover td {{ background: #f0f7ff; }}
        tr:nth-child(even) td {{ background: #fafafa; }}
        tr:nth-child(even):hover td {{ background: #f0f7ff; }}
        a {{ color: #1A73E8; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
        footer {{ text-align: center; padding: 20px; font-size: .75rem; color: #aaa; }}
        input[type=text] {{ width: calc(100% - 40px); padding: 10px 16px; font-size: .9rem;
                            border: 1px solid #ddd; border-radius: 6px; margin: 12px 20px; }}
      </style>
    </head>
    <body>
    <header>
      <h1>Reporte MercadoLibre — {nickname}</h1>
      <p>Generado el {fecha.strftime('%d/%m/%Y a las %H:%M')}</p>
    </header>

    <div class="cards">
      <div class="card"><div class="val">{len(items)}</div><div class="lbl">Publicaciones activas</div></div>
      <div class="card"><div class="val">{total_stock:,}</div><div class="lbl">Stock total disponible</div></div>
      <div class="card"><div class="val">{total_vendidos:,}</div><div class="lbl">Unidades vendidas (acumulado)</div></div>
      <div class="card"><div class="val">${precio_prom:,.0f}</div><div class="lbl">Precio promedio ARS</div></div>
      <div class="card"><div class="val">{len(cambios.get('cambios_precio',[]))}</div><div class="lbl">Cambios de precio</div></div>
    </div>

    <section>
      <h2>Ventas estimadas vs. reporte anterior</h2>
      <p style="padding:14px 20px;font-size:.9rem;">{ventas_est_str}</p>
    </section>

    <section>
      <h2>Top marcas por unidades vendidas</h2>
      <table>
        <tr><th>#</th><th>Marca</th><th>Publicaciones</th><th>Vendidos totales</th><th>Ventas Rec/Día</th></tr>
        {filas_marcas}
      </table>
    </section>

    <section>
      <h2>Publicaciones (top 300 por vendidos)</h2>
      <input type="text" id="buscar" placeholder="Filtrar por título, marca o tipo…" oninput="filtrar()">
      <table id="tabla">
        <tr>
          <th>Título</th><th>Marca</th><th>Tipo</th><th>Precio</th>
          <th>Stock</th><th>Vendidos</th><th>Rec/Día</th><th>Env. Gratis</th>
        </tr>
        {filas_items}
      </table>
    </section>

    <footer>espiar_vendedor.py — AgenciaML</footer>

    <script>
      function filtrar() {{
        const q = document.getElementById('buscar').value.toLowerCase();
        document.querySelectorAll('#tabla tr:not(:first-child)').forEach(tr => {{
          tr.style.display = tr.textContent.toLowerCase().includes(q) ? '' : 'none';
        }});
      }}
    </script>
    </body>
    </html>
    """)

    ruta.write_text(html, encoding="utf-8")
    print(f"[✓] HTML guardado:  {ruta}")
    return ruta


# ─── WhatsApp ──────────────────────────────────────────────────────────────────

def armar_mensaje(nickname: str, items: list[dict], cambios: dict, fecha: datetime) -> str:
    ts = fecha.strftime("%d/%m/%Y %H:%M")

    marcas: dict = defaultdict(int)
    for it in items:
        marcas[it["marca"]] += it["vendidos"]
    top5 = sorted(marcas.items(), key=lambda x: x[1], reverse=True)[:5]

    ventas_rec = (
        f"~{cambios['ventas_rec_dia_total']}/día en últimos {cambios['delta_dias']} días"
        if cambios.get("ventas_rec_dia_total") is not None
        else "Primer snapshot"
    )

    lineas = [
        f"*Reporte ML* — {ts}",
        f"*{nickname}*",
        f"Publicaciones: {len(items)} | Stock total: {sum((i['stock'] or 0) for i in items):,}",
        f"Ventas estimadas: {ventas_rec}",
        "",
        "*Top marcas (vendidos acum.):*",
    ]
    for m, v in top5:
        lineas.append(f"  • {m}: {v:,} uds vendidas")

    if cambios["cambios_precio"]:
        lineas.append(f"\nCambios de precio: {len(cambios['cambios_precio'])}")
        for cp in cambios["cambios_precio"][:3]:
            signo = "▲" if cp["delta%"] > 0 else "▼"
            lineas.append(f"  {signo} {cp['titulo'][:35]}… ({cp['delta%']:+.1f}%)")

    if cambios["cambios_stock"]:
        lineas.append(f"\nCambios de stock: {len(cambios['cambios_stock'])}")

    lineas.append("\n_espiar_vendedor.py_")
    return "\n".join(lineas)


def enviar_whatsapp(mensaje: str) -> None:
    if not TWILIO_DISPONIBLE:
        print("[!] Twilio no instalado — saltando WhatsApp.")
        return
    if not all([TWILIO_SID, TWILIO_TOKEN, WA_TO]):
        print("[!] Credenciales de Twilio incompletas — saltando WhatsApp.")
        return
    client = TwilioClient(TWILIO_SID, TWILIO_TOKEN)
    msg    = client.messages.create(from_=WA_FROM, to=WA_TO, body=mensaje)
    print(f"[✓] WhatsApp enviado. SID: {msg.sid}")


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Espía publicaciones de un vendedor ML Argentina.")
    parser.add_argument("nickname", help="Nickname del vendedor (ej: todoairelibregd)")
    parser.add_argument("--carpeta", default=None, help="Subcarpeta en reportes_ml/")
    args = parser.parse_args()

    nickname    = args.nickname.strip()
    fecha_ahora = datetime.now()
    subcarpeta  = args.carpeta if args.carpeta else nickname.upper()
    reports_dir = REPORTS_ROOT / subcarpeta
    reports_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n=== Espiando: {nickname} | Carpeta: {reports_dir} ===\n")

    print("[1/5] Scraping páginas de la tienda con Playwright…")
    t0 = time.time()
    raw_items = scrape_paginas_vendedor(nickname)
    print(f"      {len(raw_items)} publicaciones encontradas ({time.time()-t0:.0f}s)")

    if not raw_items:
        print("      Sin publicaciones activas. Saliendo.")
        sys.exit(0)

    # Filtrar ítems sin item_id (catálogos sin wid, no son publicaciones del vendedor)
    raw_items = [i for i in raw_items if i.get("item_id")]
    print(f"      {len(raw_items)} con ID individual (wid).")

    print(f"\n[2/5] Descargando detalle de {len(raw_items)} publicaciones ({WORKERS} hilos)…")
    t0 = time.time()
    items = enriquecer_items_paralelo(raw_items)
    print(f"      Listo en {time.time()-t0:.0f}s")

    print("\n[3/5] Cargando reporte anterior para comparar…")
    anteriores, fecha_ant = cargar_reporte_anterior(nickname, reports_dir)
    cambios = comparar_snapshots(items, anteriores, fecha_ahora, fecha_ant)

    print("\n[4/5] Guardando Excel y HTML…")
    guardar_excel(items, nickname, fecha_ahora, reports_dir, cambios)
    generar_html(items, nickname, fecha_ahora, cambios, reports_dir)

    total_vendidos = sum(i["vendidos"] for i in items)
    total_stock    = sum((i["stock"] or 0) for i in items)

    print(f"""
── Resumen ──────────────────────────────────────
   Publicaciones activas : {len(items)}
   Stock total           : {total_stock:,}
   Ventas acumuladas     : {total_vendidos:,} uds
   Ventas rec/día        : {cambios['ventas_rec_dia_total'] or 'Primer snapshot'}
   Cambios de precio     : {len(cambios['cambios_precio'])}
   Cambios de stock      : {len(cambios['cambios_stock'])}
   Nuevas publicaciones  : {len(cambios['nuevos'])}
   Eliminadas            : {len(cambios['eliminados'])}
─────────────────────────────────────────────────""")

    print("[5/5] Enviando WhatsApp…")
    enviar_whatsapp(armar_mensaje(nickname, items, cambios, fecha_ahora))

    print("\n=== Listo ===\n")


if __name__ == "__main__":
    main()
