"""
agentes/outreach.py — Agente de prospección y outreach
-------------------------------------------------------
Trigger: "Outreach [rubro]" o "Outreach [rubro] en [ciudad]"
Busca prospectos en MercadoLibre, Tiendanube y Google Maps,
encuentra su Instagram y genera un DM personalizado listo para enviar.
"""

import os
import re
import time
import subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from urllib.parse import quote_plus

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


def _leer_identidad() -> str:
    ruta = Path("agencia.md")
    if not ruta.exists():
        return ""
    lineas = [l for l in ruta.read_text(encoding="utf-8").splitlines()
              if not (l.strip().startswith("(") and l.strip().endswith(")"))]
    return "\n".join(lineas)


def _nombre_agencia() -> str:
    for linea in _leer_identidad().splitlines():
        if linea.strip() and not linea.startswith("#"):
            return linea.strip()
    return "Impulse Agency"


def _extraer_instagram(html: str) -> str:
    patron = r'(?:href=["\'])(https?://(?:www\.)?instagram\.com/(?!p/|reel/|explore/|stories/)[a-zA-Z0-9._]{3,30}/?)["\']'
    match = re.search(patron, html, re.IGNORECASE)
    if match:
        return match.group(1).rstrip("/").split("?")[0]
    patron2 = r'instagram\.com/(?!p/|reel/|explore/|stories/)([a-zA-Z0-9._]{3,30})'
    match2 = re.search(patron2, html, re.IGNORECASE)
    if match2:
        return f"https://www.instagram.com/{match2.group(1)}"
    return ""


def _scrape(url: str, espera: float = 1.5) -> str:
    from playwright.sync_api import sync_playwright
    html = ""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        try:
            page.goto(url, timeout=18000, wait_until="domcontentloaded")
            time.sleep(espera)
            html = page.content()
        except Exception:
            pass
        finally:
            browser.close()
    return html


def _instagram_via_buscador(nombre: str) -> str:
    query = quote_plus(f'"{nombre}" instagram.com')
    for base_url in [
        f"https://search.yahoo.com/search?p={query}",
        f"https://www.bing.com/search?q={query}&count=5",
    ]:
        html = _scrape(base_url, espera=2.0)
        ig = _extraer_instagram(html)
        if ig:
            return ig
    return ""


def _instagram_via_google(nombre: str) -> str:
    return _instagram_via_buscador(nombre)


# ─── Fuente 1: MercadoLibre ───────────────────────────────────────────────────

def _meli_slug(rubro: str) -> str:
    slug = re.sub(r'[^a-z0-9]+', '-', rubro.lower()).strip('-')
    return slug


def buscar_meli(rubro: str, limite: int = 5) -> list[dict]:
    from playwright.sync_api import sync_playwright

    print(f"  [MercadoLibre] Buscando vendedores de '{rubro}'...")
    prospectos = []
    vistos = set()

    # Fase 1: scraping dentro del contexto Playwright
    datos_raw = []
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled", "--disable-dev-shm-usage"],
        )
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800},
            locale="es-AR",
        )
        ctx.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        page = ctx.new_page()

        try:
            slug = _meli_slug(rubro)
            listing_url = f"https://listado.mercadolibre.com.ar/{slug}"
            page.goto(listing_url, timeout=25000, wait_until="networkidle")
            time.sleep(4)

            html_listing = page.content()
            if len(html_listing) < 50000:
                page.goto(
                    f"https://listado.mercadolibre.com.ar/_NoIndex_True?q={quote_plus(rubro)}",
                    timeout=25000, wait_until="networkidle"
                )
                time.sleep(4)
                html_listing = page.content()

            links_articulos = list(dict.fromkeys([
                l.split('#')[0].split('?')[0]
                for l in re.findall(r'https://articulo\.mercadolibre\.com\.ar/MLA[^"\'>\s&]+', html_listing)
            ]))
            print(f"    {len(links_articulos)} artículos encontrados en listado")

            for link in links_articulos[:limite * 4]:
                if len(datos_raw) >= limite:
                    break
                try:
                    page.goto(link, timeout=25000, wait_until="load")
                    time.sleep(3)
                    html_prod = page.content()

                    m_id = re.search(r'"seller_id"\s*:\s*(\d+)', html_prod)
                    m_shop = re.search(r'"shop_name"\s*:\s*"([^"]+)"', html_prod)

                    if not m_id:
                        continue
                    seller_id = m_id.group(1)
                    if seller_id in vistos:
                        continue
                    vistos.add(seller_id)

                    nombre = m_shop.group(1).title() if m_shop else seller_id
                    tiene_redes = bool(re.search(r'instagram|tiktok|youtube|facebook', html_prod, re.IGNORECASE))

                    datos_raw.append({
                        "nombre": nombre,
                        "seller_id": seller_id,
                        "tiene_redes": tiene_redes,
                    })
                    print(f"    ✓ Encontrado: {nombre}")

                except Exception as e:
                    print(f"    ✗ artículo omitido: {type(e).__name__}: {e}")
                    continue

        except Exception as e:
            print(f"    Error MercadoLibre: {e}")
        finally:
            browser.close()

    # Fase 2: buscar Instagram fuera del contexto Playwright (evita anidamiento)
    for d in datos_raw:
        instagram = _instagram_via_google(d["nombre"])
        dolor = (
            "vende en MercadoLibre con redes pero sin estrategia de contenido ni ads"
            if d["tiene_redes"]
            else "vende solo en MercadoLibre, sin redes sociales activas ni marca propia"
        )
        prospectos.append({
            "nombre": d["nombre"],
            "plataforma": "MercadoLibre",
            "rubro": rubro,
            "url_perfil": f"https://www.mercadolibre.com.ar/perfil/{d['seller_id']}",
            "instagram": instagram,
            "dolor": dolor,
        })
        print(f"    IG {d['nombre']}: {instagram or '(no encontrado)'}")

    return prospectos


# ─── Fuente 2: Tiendanube ─────────────────────────────────────────────────────

def buscar_tiendanube(rubro: str, limite: int = 5) -> list[dict]:
    from playwright.sync_api import sync_playwright

    print(f"  [Tiendanube] Buscando tiendas de '{rubro}'...")
    prospectos = []

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled"],
        )
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="es-AR",
        )
        ctx.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        page = ctx.new_page()

        try:
            # Buscar en el directorio público de Tiendanube
            search_url = f"https://www.tiendanube.com/tiendas-online?q={quote_plus(rubro)}"
            page.goto(search_url, timeout=20000, wait_until="domcontentloaded")
            time.sleep(3)
            html = page.content()

            # Extraer URLs de tiendas .mitiendanube.com
            urls = re.findall(r'https?://[a-zA-Z0-9\-]+\.mitiendanube\.com', html)
            urls_base = list(dict.fromkeys(urls))

            # Si el directorio no devuelve resultados, buscar via Yahoo y Bing
            if not urls_base:
                for buscador_url in [
                    f"https://search.yahoo.com/search?p=site%3Amitiendanube.com+{quote_plus(rubro)}",
                    f"https://www.bing.com/search?q=site%3Amitiendanube.com+{quote_plus(rubro)}&count=20",
                ]:
                    page.goto(buscador_url, timeout=20000, wait_until="domcontentloaded")
                    time.sleep(3)
                    html_b = page.content()
                    urls_b = re.findall(r'https?://[a-zA-Z0-9\-]+\.mitiendanube\.com', html_b)
                    urls_base = list(dict.fromkeys(urls_b))
                    if urls_base:
                        print(f"    Buscador encontró {len(urls_base)} tiendas")
                        break
                if not urls_base:
                    print(f"    Sin resultados en buscadores")

            for store_url in urls_base[:limite * 2]:
                if len(prospectos) >= limite:
                    break
                try:
                    page.goto(store_url, timeout=15000, wait_until="domcontentloaded")
                    time.sleep(1.5)
                    store_html = page.content()
                    if not store_html:
                        continue

                    m_titulo = re.search(r'<title>([^<|–\-]+)', store_html)
                    nombre = m_titulo.group(1).strip() if m_titulo else store_url

                    instagram = _extraer_instagram(store_html)
                    if not instagram:
                        instagram = _instagram_via_google(nombre)

                    prospectos.append({
                        "nombre": nombre,
                        "plataforma": "Tiendanube",
                        "rubro": rubro,
                        "url_perfil": store_url,
                        "instagram": instagram,
                        "dolor": "tiene tienda online pero probablemente no invierte en Ads ni en optimización de conversión",
                    })
                    print(f"    ✓ {nombre} — IG: {instagram or '(no encontrado)'}")

                except Exception as e:
                    print(f"    ✗ tienda omitida: {type(e).__name__}")
                    continue

        except Exception as e:
            print(f"    Error Tiendanube: {e}")
        finally:
            browser.close()

    return prospectos


# ─── Fuente 3: Google Maps ────────────────────────────────────────────────────

def buscar_google_maps(rubro: str, ciudad: str = "Buenos Aires", limite: int = 5) -> list[dict]:
    from playwright.sync_api import sync_playwright

    print(f"  [Google Maps] Buscando '{rubro}' en {ciudad}...")
    nombres_raw = []

    # Fase 1: extraer nombres de negocios desde los links de Maps (sin hacer clicks)
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled"],
        )
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800},
            locale="es-AR",
        )
        ctx.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        page = ctx.new_page()

        try:
            query = quote_plus(f"{rubro} {ciudad}")
            page.goto(f"https://www.google.com/maps/search/{query}", timeout=20000, wait_until="domcontentloaded")
            time.sleep(4)

            # Extraer nombres de negocios desde los hrefs de maps/place (están URL-encoded en el path)
            html = page.content()
            place_links = re.findall(r'href="https://www\.google\.com/maps/place/([^/!?]+)', html)
            vistos = set()
            for raw in place_links:
                from urllib.parse import unquote_plus
                nombre = unquote_plus(raw.replace('+', ' ')).strip()
                if nombre and nombre not in vistos and len(nombre) > 3:
                    vistos.add(nombre)
                    nombres_raw.append(nombre)
                if len(nombres_raw) >= limite * 2:
                    break

            print(f"    {len(nombres_raw)} negocios encontrados en Maps")

        except Exception as e:
            print(f"    Error Google Maps: {e}")
        finally:
            browser.close()

    # Fase 2: buscar Instagram para cada negocio (fuera del contexto Playwright)
    prospectos = []
    for nombre in nombres_raw[:limite]:
        instagram = _instagram_via_google(nombre)
        prospectos.append({
            "nombre": nombre,
            "plataforma": "Google Maps",
            "rubro": rubro,
            "url_perfil": "",
            "instagram": instagram,
            "dolor": "negocio físico con presencia en Maps pero sin estrategia digital ni MercadoLibre optimizado",
        })
        print(f"    ✓ {nombre} — IG: {instagram or '(no encontrado)'}")

    return prospectos


# ─── Generación de mensajes DM ────────────────────────────────────────────────

def generar_mensaje(prospecto: dict, identidad: str, client=None) -> str:
    import anthropic

    if client is None:
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

    contexto = f"\nIDENTIDAD Y TONO DE LA AGENCIA:\n{identidad}\n" if identidad.strip() else ""

    prompt = f"""Sos consultor senior de e-commerce de la agencia "{_nombre_agencia()}" en Argentina.
Escribí un DM de Instagram corto y personalizado para abrir conversación con este prospecto.
{contexto}
PROSPECTO:
- Nombre: {prospecto['nombre']}
- Vende en: {prospecto['plataforma']}
- Rubro: {prospecto['rubro']}
- Problema detectado: {prospecto['dolor']}

ESTRUCTURA DEL MENSAJE:
1. Apertura específica que demuestre que miraste su negocio (mencioná su rubro o plataforma; prohibido "vi tu perfil" o "me encanta tu marca")
2. El problema detectado, planteado como observación útil, no como crítica
3. Una pista de valor concreta: qué podría lograr si lo resuelve
4. Cierre con pregunta abierta que invite a responder (prohibido "¿Te interesa?" o "¿Hablamos?")

REGLAS:
- Máximo 4 líneas
- Tono: consultor que aporta valor, no vendedor que persigue
- Máximo 2 emojis
- Español argentino (vos, tenés, podés)
- Sin autopresentaciones largas ni saludos genéricos

Devolvé SOLO el texto del mensaje, sin comillas ni explicación."""

    msg = client.messages.create(
        model="claude-haiku-4-5",
        max_tokens=300,
        messages=[{"role": "user", "content": prompt}]
    )
    return msg.content[0].text.strip()


# ─── Exportadores ─────────────────────────────────────────────────────────────

def exportar_excel(prospectos: list[dict], ruta: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospectos"

    headers = ["Nombre", "Plataforma", "Rubro", "URL / Perfil", "Instagram", "Dolor detectado", "Mensaje DM"]
    fill_header = PatternFill("solid", fgColor="1AE82F")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="000000")
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")

    for i, p in enumerate(prospectos, 2):
        ws.cell(row=i, column=1, value=p["nombre"])
        ws.cell(row=i, column=2, value=p["plataforma"])
        ws.cell(row=i, column=3, value=p["rubro"])
        ws.cell(row=i, column=4, value=p["url_perfil"])
        ig = ws.cell(row=i, column=5, value=p["instagram"])
        if p["instagram"]:
            ig.hyperlink = p["instagram"]
            ig.font = Font(color="22CFFF", underline="single")
        ws.cell(row=i, column=6, value=p["dolor"])
        ws.cell(row=i, column=7, value=p.get("mensaje_dm", ""))

    for col, ancho in enumerate([30, 16, 20, 42, 38, 42, 65], 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = ancho

    wb.save(ruta)


def exportar_html(prospectos: list[dict], ruta: Path, rubro: str, ciudad: str) -> None:
    con_ig = sum(1 for p in prospectos if p["instagram"])
    con_msg = sum(1 for p in prospectos if p.get("mensaje_dm"))

    filas = ""
    for p in prospectos:
        badge_class = p["plataforma"].lower().replace(" ", "")
        ig_html = (
            f'<a href="{p["instagram"]}" target="_blank">@{p["instagram"].split("/")[-1]}</a>'
            if p["instagram"] else '<span class="na">—</span>'
        )
        mensaje = (p.get("mensaje_dm") or "").replace("\n", "<br>")
        filas += f"""
        <tr>
          <td><strong>{p['nombre']}</strong><br><small><a href="{p['url_perfil']}" target="_blank">{p['url_perfil'][:45]}...</a></small></td>
          <td><span class="badge {badge_class}">{p['plataforma']}</span></td>
          <td>{p['rubro']}</td>
          <td>{ig_html}</td>
          <td class="dolor">{p['dolor']}</td>
          <td class="mensaje">{mensaje}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Outreach — {rubro}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0d0d0d;color:#f0f0f0;padding:2rem}}
  h1{{font-size:1.6rem;margin-bottom:.3rem;color:#1ae82f}}
  .meta{{color:#888;font-size:.85rem;margin-bottom:1.5rem}}
  .stats{{display:flex;gap:1rem;margin-bottom:1.5rem;flex-wrap:wrap}}
  .stat{{background:#1a1a1a;border-radius:8px;padding:1rem 1.5rem;border-left:3px solid #1ae82f}}
  .stat-num{{font-size:1.8rem;font-weight:700;color:#1ae82f}}
  .stat-label{{font-size:.78rem;color:#888;margin-top:.2rem}}
  table{{width:100%;border-collapse:collapse;background:#1a1a1a;border-radius:10px;overflow:hidden}}
  th{{background:#1ae82f;color:#000;padding:.75rem 1rem;text-align:left;font-size:.8rem;text-transform:uppercase;letter-spacing:.05em}}
  td{{padding:.85rem 1rem;border-bottom:1px solid #2a2a2a;font-size:.88rem;vertical-align:top}}
  tr:hover td{{background:#222}}
  a{{color:#22cfff;text-decoration:none}}
  a:hover{{text-decoration:underline}}
  .badge{{padding:2px 9px;border-radius:20px;font-size:.72rem;font-weight:700}}
  .mercadolibre{{background:#ffe600;color:#000}}
  .tiendanube{{background:#22cfff;color:#000}}
  .googlemaps{{background:#4285f4;color:#fff}}
  .dolor{{color:#aaa;font-size:.82rem}}
  .mensaje{{font-size:.83rem;line-height:1.55;color:#ddd;min-width:220px}}
  .na{{color:#555}}
  small a{{color:#555;font-size:.75rem}}
</style>
</head>
<body>
  <h1>Outreach — {rubro}</h1>
  <p class="meta">{_nombre_agencia()} · {ciudad} · {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
  <div class="stats">
    <div class="stat"><div class="stat-num">{len(prospectos)}</div><div class="stat-label">Prospectos</div></div>
    <div class="stat"><div class="stat-num">{con_ig}</div><div class="stat-label">Con Instagram</div></div>
    <div class="stat"><div class="stat-num">{con_msg}</div><div class="stat-label">Mensajes listos</div></div>
  </div>
  <table>
    <thead><tr><th>Negocio</th><th>Plataforma</th><th>Rubro</th><th>Instagram</th><th>Dolor detectado</th><th>Mensaje DM</th></tr></thead>
    <tbody>{filas}</tbody>
  </table>
</body>
</html>"""
    ruta.write_text(html, encoding="utf-8")


# ─── Función principal ────────────────────────────────────────────────────────

def run(
    rubros: list[str],
    ciudad: str = "Buenos Aires",
    limite_por_fuente: int = 5,
    generar_mensajes: bool = True,
) -> Path:
    identidad = _leer_identidad()
    fecha = datetime.now().strftime("%Y%m%d")
    slug_combinado = "-".join(re.sub(r'[^a-z0-9]+', '-', r.lower()).strip('-') for r in rubros)

    carpeta = Path("reportes_ml/AgenciaML")
    carpeta.mkdir(parents=True, exist_ok=True)

    print(f"\n[Outreach] Rubros: {', '.join(rubros)} | Ciudad: {ciudad} | Límite por fuente: {limite_por_fuente}")
    print("=" * 55)

    prospectos = []
    for rubro in rubros:
        print(f"\n── Rubro: {rubro.upper()} ──")
        prospectos += buscar_tiendanube(rubro, limite_por_fuente)
        prospectos += buscar_google_maps(rubro, ciudad, limite_por_fuente)

    if generar_mensajes and prospectos:
        import anthropic

        print(f"\n[Claude] Generando mensajes DM para {len(prospectos)} prospectos en paralelo...")
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        with ThreadPoolExecutor(max_workers=5) as pool:
            futuros = {pool.submit(generar_mensaje, p, identidad, client): p for p in prospectos}
            for futuro in as_completed(futuros):
                p = futuros[futuro]
                try:
                    p["mensaje_dm"] = futuro.result()
                    print(f"  ✓ {p['nombre']}")
                except Exception as e:
                    p["mensaje_dm"] = ""
                    print(f"  ✗ {p['nombre']}: {e}")

    titulo = " · ".join(rubros)
    ruta_excel = carpeta / f"outreach_{slug_combinado}_{fecha}.xlsx"
    ruta_html = carpeta / f"outreach_{slug_combinado}_{fecha}.html"

    exportar_excel(prospectos, ruta_excel)
    exportar_html(prospectos, ruta_html, titulo, ciudad)

    print(f"\n✅ {len(prospectos)} prospectos · {sum(1 for p in prospectos if p['instagram'])} con Instagram")
    print(f"   Excel : {ruta_excel}")
    print(f"   HTML  : {ruta_html}")

    subprocess.Popen(["open", str(ruta_html)])
    subprocess.Popen(["afplay", "/System/Library/Sounds/Glass.aiff"])

    return ruta_html


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Agente de outreach para Impulse Agency")
    parser.add_argument("rubros", nargs="+", help="Rubros a buscar (ej: indumentaria electronica)")
    parser.add_argument("--ciudad", default="Buenos Aires", help="Ciudad para Google Maps")
    parser.add_argument("--limite", type=int, default=5, help="Prospectos por fuente")
    parser.add_argument("--sin-mensajes", action="store_true", help="No generar DMs con Claude")
    args = parser.parse_args()

    run(args.rubros, args.ciudad, args.limite, not args.sin_mensajes)
