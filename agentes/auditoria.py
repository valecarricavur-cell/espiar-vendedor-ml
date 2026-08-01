"""
agentes/auditoria.py — Auditoría de cuenta propia en Mercado Libre
-------------------------------------------------------------------
Trigger: "Auditoria [nickname]"
Lee el Excel del vendedor y genera un reporte de diagnóstico profesional.
"""

import sys
import re
import glob
import subprocess
from datetime import datetime
from pathlib import Path


def _leer_identidad() -> str:
    ruta = Path("agencia.md")
    if not ruta.exists():
        return ""
    lineas = [l for l in ruta.read_text(encoding="utf-8").splitlines()
              if not (l.strip().startswith("(") and l.strip().endswith(")"))]
    return "\n".join(lineas)


def _nombre_agencia() -> str:
    for linea in _leer_identidad().splitlines():
        if linea.strip() and not linea.startswith("#"):
            return linea.strip()
    return "Impulse Agency"


def leer_datos_excel(nickname: str, carpeta: str = None) -> dict:
    """Lee el último Excel disponible para el nickname dado."""
    import openpyxl

    if carpeta:
        rutas = sorted(glob.glob(f"reportes_ml/{carpeta.upper()}/{nickname}_*.xlsx"))
    else:
        rutas = sorted(glob.glob(f"reportes_ml/**/{nickname}_*.xlsx", recursive=True))

    if not rutas:
        return {}

    wb = openpyxl.load_workbook(rutas[-1])
    ws = wb.active
    headers = [c.value for c in ws[1]]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or str(row[0]).startswith("▼") or row[0] == "ID":
            continue
        items.append(dict(zip(headers, row)))

    def safe_int(v):
        try: return int(v or 0)
        except: return 0

    def safe_float(v):
        try: return float(v or 0)
        except: return 0.0

    vendidos = [safe_int(i.get("Vendidos Totales")) for i in items]
    stocks   = [safe_int(i.get("Stock Total")) for i in items]
    precios  = [safe_float(i.get("Precio (ARS)")) for i in items]

    sin_ventas   = [i for i, v in zip(items, vendidos) if v == 0]
    stock_bajo   = [i for i, s in zip(items, stocks)   if 0 < s <= 3]
    sin_stock    = [i for i, s in zip(items, stocks)   if s == 0]
    top_vendidos = sorted(items, key=lambda x: safe_int(x.get("Vendidos Totales")), reverse=True)[:10]

    return {
        "archivo":       Path(rutas[-1]).name,
        "total":         len(items),
        "total_vendidos": sum(vendidos),
        "precio_prom":   round(sum(precios) / max(len(precios), 1)),
        "sin_ventas":    sin_ventas[:10],
        "sin_ventas_n":  len(sin_ventas),
        "stock_bajo":    stock_bajo[:10],
        "stock_bajo_n":  len(stock_bajo),
        "sin_stock":     sin_stock[:5],
        "sin_stock_n":   len(sin_stock),
        "top_vendidos":  top_vendidos,
    }


def _md_a_html(texto: str) -> str:
    lineas = []
    for linea in texto.split("\n"):
        l = linea.strip()
        if not l:
            lineas.append("<br>")
        elif l.startswith("### "):
            lineas.append(f'<h3>{l[4:]}</h3>')
        elif l.startswith("## "):
            lineas.append(f'<h2>{l[3:]}</h2>')
        elif l.startswith("# "):
            lineas.append(f'<h1>{l[2:]}</h1>')
        elif l.startswith("---"):
            lineas.append('<hr>')
        elif l.startswith("- ") or l.startswith("• "):
            lineas.append(f'<li>{l[2:]}</li>')
        elif re.match(r'^\d+\.', l):
            lineas.append(f'<li>{l}</li>')
        elif l.startswith("> "):
            lineas.append(f'<blockquote>{l[2:]}</blockquote>')
        else:
            l = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', l)
            l = re.sub(r'\*(.+?)\*', r'<em>\1</em>', l)
            lineas.append(f'<p>{l}</p>')
    return "\n".join(lineas)


def guardar_doc(nickname: str, contenido: str, datos: dict, fecha: datetime) -> Path:
    carpeta = Path("reportes_ml") / "AgenciaML"
    carpeta.mkdir(parents=True, exist_ok=True)
    ts   = fecha.strftime("%Y%m%d_%H%M%S")
    ruta = carpeta / f"auditoria_{nickname}_{ts}.html"

    nombre = _nombre_agencia()
    cuerpo = _md_a_html(contenido)

    d = datos
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Auditoría ML — {nickname}</title>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    *{{box-sizing:border-box;margin:0;padding:0}}
    body{{font-family:'Inter',sans-serif;background:#f8f9fa;color:#1a1a1a;line-height:1.75;padding:40px 20px 80px}}
    .doc{{background:#fff;max-width:820px;margin:0 auto;padding:60px 72px;border-radius:4px;box-shadow:0 1px 3px rgba(0,0,0,.12),0 4px 20px rgba(0,0,0,.06)}}
    .doc-header{{border-bottom:1px solid #e8eaed;padding-bottom:24px;margin-bottom:36px}}
    .doc-tag{{display:inline-block;background:#1ae82f22;color:#0d7a1a;font-size:11px;font-weight:600;letter-spacing:.06em;text-transform:uppercase;padding:4px 10px;border-radius:4px;margin-bottom:12px}}
    .doc-title{{font-size:26px;font-weight:700;margin-bottom:8px}}
    .doc-meta{{font-size:13px;color:#80868b}}
    .stats{{display:flex;gap:12px;flex-wrap:wrap;margin:24px 0}}
    .stat{{background:#f8f9fa;border-radius:8px;padding:14px 18px;flex:1;min-width:120px;text-align:center;border:1px solid #e8eaed}}
    .stat .val{{font-size:22px;font-weight:700;color:#1a1a1a}}
    .stat .lbl{{font-size:11px;color:#80868b;margin-top:2px;text-transform:uppercase;letter-spacing:.04em}}
    .stat.alerta .val{{color:#d93025}}
    .stat.warn .val{{color:#f29900}}
    h1{{font-size:20px;font-weight:700;margin:32px 0 12px}}
    h2{{font-size:17px;font-weight:600;margin:28px 0 10px}}
    h3{{font-size:13px;font-weight:600;margin:20px 0 6px;text-transform:uppercase;letter-spacing:.05em;color:#3c4043}}
    p{{font-size:15px;color:#3c4043;margin:6px 0}}
    li{{font-size:15px;color:#3c4043;margin:5px 0 5px 20px;list-style:disc}}
    blockquote{{background:#fff8e1;border-left:3px solid #f29900;padding:14px 20px;margin:12px 0;border-radius:0 6px 6px 0;font-size:15px;font-weight:500;color:#1a1a1a}}
    hr{{border:none;border-top:1px solid #e8eaed;margin:28px 0}}
    strong{{color:#1a1a1a;font-weight:600}}
    .doc-footer{{margin-top:48px;padding-top:20px;border-top:1px solid #e8eaed;font-size:12px;color:#9aa0a6;display:flex;justify-content:space-between}}
  </style>
</head>
<body>
<div class="doc">
  <div class="doc-header">
    <div class="doc-tag">Auditoría Mercado Libre</div>
    <div class="doc-title">{nickname}</div>
    <div class="doc-meta">{nombre} &nbsp;·&nbsp; {fecha.strftime('%d de %B de %Y, %H:%M')}</div>
  </div>

  <div class="stats">
    <div class="stat"><div class="val">{d.get('total', 0)}</div><div class="lbl">Publicaciones</div></div>
    <div class="stat"><div class="val">{d.get('total_vendidos', 0):,}</div><div class="lbl">Vendidos</div></div>
    <div class="stat"><div class="val">${d.get('precio_prom', 0):,}</div><div class="lbl">Precio prom.</div></div>
    <div class="stat alerta"><div class="val">{d.get('sin_ventas_n', 0)}</div><div class="lbl">Sin ventas</div></div>
    <div class="stat warn"><div class="val">{d.get('stock_bajo_n', 0)}</div><div class="lbl">Stock bajo</div></div>
    <div class="stat alerta"><div class="val">{d.get('sin_stock_n', 0)}</div><div class="lbl">Sin stock</div></div>
  </div>

  {cuerpo}

  <div class="doc-footer">
    <span>{nombre}</span>
    <span>Auditoría · {d.get('archivo', '')}</span>
  </div>
</div>
</body>
</html>"""

    ruta.write_text(html, encoding="utf-8")
    return ruta


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("nickname")
    parser.add_argument("--carpeta", default=None)
    parser.add_argument("--archivo", help="Archivo .md con el análisis generado")
    args = parser.parse_args()

    datos = leer_datos_excel(args.nickname, args.carpeta)
    if not datos:
        print(f"No se encontró Excel para '{args.nickname}'. Corré primero espiar_vendedor.py.")
        sys.exit(1)

    if args.archivo:
        contenido = Path(args.archivo).read_text(encoding="utf-8")
    else:
        contenido = sys.stdin.read()

    fecha = datetime.now()
    ruta  = guardar_doc(args.nickname, contenido, datos, fecha)
    subprocess.run(["open", str(ruta)], check=False)
    subprocess.run(
        ["osascript", "-e",
         f'display notification "Auditoría lista — {args.nickname}" with title "{_nombre_agencia()}" sound name "Glass"'],
        check=False, capture_output=True
    )
    print(f"Guardado: {ruta}")


if __name__ == "__main__":
    main()
