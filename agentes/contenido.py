"""
agentes/contenido.py — Agente de Contenido e Ideas
---------------------------------------------------
Lee el último reporte de competencia y usa Claude para generar:
  - Análisis del mercado (qué vende, qué tendencias hay)
  - 5 ideas de posts para Instagram con caption listo para publicar
  - Oportunidades de precio detectadas vs la competencia
  - Sugerencias de productos a agregar al catálogo

Salida: HTML + Markdown en reportes_ml/{cliente}/contenido_{fecha}.html
"""

import os
import json
import glob
import textwrap
from datetime import datetime
from pathlib import Path

import anthropic
import openpyxl

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# ─── Leer datos del último reporte ───────────────────────────────────────────

def _leer_ultimo_reporte(carpeta_reportes: Path, nickname: str) -> dict:
    """Extrae resumen del último Excel generado por el agente de espionaje."""
    archivos = sorted(glob.glob(str(carpeta_reportes / f"{nickname}_*.xlsx")))
    if not archivos:
        return {}

    ruta = archivos[-1]
    wb = openpyxl.load_workbook(ruta)

    # Hoja 1: publicaciones
    ws = wb.active
    items = []
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Saltar filas vacías, separadores y encabezados repetidos (sección watchlist)
        if not row[0] or str(row[0]).startswith("▼") or row[0] == "ID":
            continue
        items.append(dict(zip(headers, row)))

    # Hoja 2: por marca
    marcas = []
    if "Por Marca" in wb.sheetnames:
        ws2 = wb["Por Marca"]
        h2 = [c.value for c in ws2[1]]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0]:
                marcas.append(dict(zip(h2, row)))

    # Hoja 3: por tipo
    tipos = []
    if "Por Tipo" in wb.sheetnames:
        ws3 = wb["Por Tipo"]
        h3 = [c.value for c in ws3[1]]
        for row in ws3.iter_rows(min_row=2, values_only=True):
            if row[0]:
                tipos.append(dict(zip(h3, row)))

    # Top 15 por vendidos
    top = sorted(items, key=lambda x: int(x.get("Vendidos Totales") or 0) if str(x.get("Vendidos Totales", "0")).isdigit() else 0, reverse=True)[:15]

    return {
        "archivo":         Path(ruta).name,
        "total_items":     len(items),
        "total_stock":     sum(int(i.get("Stock Total") or 0) for i in items if str(i.get("Stock Total", "0")).isdigit()),
        "total_vendidos":  sum(int(i.get("Vendidos Totales") or 0) for i in items if str(i.get("Vendidos Totales", "0")).isdigit()),
        "precio_prom":     round(sum(float(i.get("Precio (ARS)") or 0) for i in items) / max(len(items), 1)),
        "top_productos":   [
            {
                "titulo":   t.get("Título", "")[:60],
                "marca":    t.get("Marca", ""),
                "tipo":     t.get("Tipo de Producto", ""),
                "precio":   t.get("Precio (ARS)"),
                "stock":    t.get("Stock Total"),
                "vendidos": t.get("Vendidos Totales"),
            }
            for t in top
        ],
        "top_marcas": marcas[:8],
        "top_tipos":  tipos[:8],
    }


# ─── Llamadas a Claude ────────────────────────────────────────────────────────

def _prompt_analisis(datos: dict, nombre_cliente: str) -> str:
    return f"""Sos un analista de e-commerce especializado en MercadoLibre Argentina.
Tu cliente es "{nombre_cliente}".

Te paso datos de la competencia que monitoreamos:

RESUMEN GENERAL:
- Publicaciones activas del competidor: {datos['total_items']}
- Stock total disponible: {datos['total_stock']}
- Unidades vendidas (acumulado): {datos['total_vendidos']}
- Precio promedio: ${datos['precio_prom']:,}

TOP 15 PRODUCTOS MÁS VENDIDOS:
{json.dumps(datos['top_productos'], ensure_ascii=False, indent=2)}

TOP MARCAS:
{json.dumps(datos['top_marcas'], ensure_ascii=False, indent=2)}

TOP TIPOS DE PRODUCTO:
{json.dumps(datos['top_tipos'], ensure_ascii=False, indent=2)}

Generá un análisis completo en español con estas secciones:
1. **Resumen ejecutivo** (3-4 oraciones sobre el estado general del competidor)
2. **Categorías estrella** (qué rubros dominan y por qué importan)
3. **Oportunidades de precio** (donde el competidor tiene precios altos y se puede competir)
4. **Productos con poco stock** (posibles quiebres de stock = oportunidad)
5. **Tendencias detectadas** (qué está creciendo o llama la atención)
6. **Recomendación principal** (una acción concreta para mi cliente esta semana)

Sé directo, práctico y usa datos concretos del análisis."""


def _prompt_ideas_instagram(datos: dict, nombre_agencia: str) -> str:
    return f"""Sos un experto en marketing digital para e-commerce en Argentina.
Trabajás para la agencia "{nombre_agencia}".

Basándote en estos datos de un competidor de MercadoLibre:
- Categorías más vendidas: {[t['Tipo de Producto'] for t in datos['top_tipos'][:5]] if datos.get('top_tipos') else []}
- Marcas top: {[m['Marca'] for m in datos['top_marcas'][:5]] if datos.get('top_marcas') else []}
- Productos estrella: {[p['titulo'] for p in datos['top_productos'][:5]]}

Creá 5 ideas de posts para Instagram, cada una con:
- **Idea**: descripción del concepto visual en 1 línea
- **Caption**: texto listo para copiar y pegar (con emojis, natural, en español argentino)
- **Hashtags**: 8-10 hashtags relevantes
- **Objetivo**: qué busca lograr este post (venta, engagement, branding)

Los posts deben ser para una tienda de e-commerce deportes/outdoor en Argentina.
Variá los formatos: carrusel, reels, historia, post estático."""


def _prompt_plan_accion(datos: dict, nombre_cliente: str) -> str:
    return f"""Sos un consultor de e-commerce para MercadoLibre Argentina.

Analizando la competencia de "{nombre_cliente}" con {datos['total_items']} publicaciones
y {datos['total_vendidos']} unidades vendidas, creá un plan de acción para esta semana:

TOP PRODUCTOS COMPETENCIA:
{json.dumps(datos['top_productos'][:8], ensure_ascii=False, indent=2)}

Generá:
1. **3 acciones inmediatas** (hacer hoy/mañana)
2. **2 acciones esta semana** (requieren más tiempo)
3. **1 experimento a probar** (algo nuevo para testear)

Para cada acción: qué hacer, por qué, y cómo medir el resultado.
Sé específico para MercadoLibre Argentina."""


def _llamar_claude(prompt: str, max_tokens: int = 1500) -> str:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return "⚠️ Falta ANTHROPIC_API_KEY en el archivo .env"

    client = anthropic.Anthropic(api_key=api_key)
    mensaje = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}],
    )
    return mensaje.content[0].text


# ─── Generar HTML de salida ───────────────────────────────────────────────────

def _md_a_html(texto: str) -> str:
    """Convierte markdown básico a HTML."""
    import re
    texto = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', texto)
    texto = re.sub(r'\*(.+?)\*', r'<em>\1</em>', texto)
    lineas = []
    for linea in texto.split('\n'):
        linea = linea.strip()
        if linea.startswith('# '):
            lineas.append(f'<h2>{linea[2:]}</h2>')
        elif linea.startswith('## '):
            lineas.append(f'<h3>{linea[3:]}</h3>')
        elif linea.startswith('- ') or linea.startswith('• '):
            lineas.append(f'<li>{linea[2:]}</li>')
        elif linea.startswith(('1.', '2.', '3.', '4.', '5.', '6.')):
            lineas.append(f'<li>{linea}</li>')
        elif linea == '':
            lineas.append('<br>')
        else:
            lineas.append(f'<p>{linea}</p>')
    return '\n'.join(lineas)


def _guardar_html(
    analisis: str,
    ideas: str,
    plan: str,
    nombre_cliente: str,
    datos: dict,
    carpeta: Path,
    fecha: datetime,
) -> Path:
    ts   = fecha.strftime("%Y%m%d_%H%M%S")
    ruta = carpeta / f"contenido_{ts}.html"

    html = textwrap.dedent(f"""<!DOCTYPE html>
    <html lang="es">
    <head>
      <meta charset="UTF-8">
      <title>Contenido IA — {nombre_cliente}</title>
      <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: 'Segoe UI', sans-serif; background: #f0f2f5; color: #333; line-height: 1.6; }}
        header {{ background: linear-gradient(135deg, #1A73E8, #0F9D58); color: white; padding: 24px 40px; }}
        header h1 {{ font-size: 1.5rem; }}
        header p  {{ opacity: .85; font-size: .9rem; margin-top: 4px; }}
        .container {{ max-width: 960px; margin: 0 auto; padding: 32px 24px; }}
        .card {{ background: white; border-radius: 12px; padding: 28px 32px;
                 box-shadow: 0 2px 8px rgba(0,0,0,.08); margin-bottom: 24px; }}
        .card h2 {{ font-size: 1.1rem; color: #1A73E8; margin-bottom: 16px;
                    padding-bottom: 10px; border-bottom: 2px solid #e8f0fe; }}
        .card h3 {{ font-size: 1rem; color: #0F9D58; margin: 16px 0 8px; }}
        .card p  {{ margin: 8px 0; color: #444; }}
        .card li {{ margin: 6px 0 6px 20px; color: #444; }}
        .card strong {{ color: #1A73E8; }}
        .stats {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 24px; }}
        .stat {{ background: white; border-radius: 10px; padding: 16px 20px;
                 box-shadow: 0 2px 8px rgba(0,0,0,.08); flex: 1; min-width: 140px; text-align: center; }}
        .stat .val {{ font-size: 1.6rem; font-weight: 700; color: #1A73E8; }}
        .stat .lbl {{ font-size: .75rem; color: #888; margin-top: 4px; }}
        .badge {{ display: inline-block; background: #e8f0fe; color: #1A73E8;
                  padding: 3px 10px; border-radius: 12px; font-size: .8rem; margin: 2px; }}
        footer {{ text-align: center; padding: 20px; font-size: .75rem; color: #aaa; }}
      </style>
    </head>
    <body>
    <header>
      <h1>Contenido IA — {nombre_cliente}</h1>
      <p>Generado el {fecha.strftime('%d/%m/%Y a las %H:%M')} · Basado en: {datos.get('archivo', '')}</p>
    </header>
    <div class="container">

    <div class="stats">
      <div class="stat"><div class="val">{datos['total_items']}</div><div class="lbl">Publicaciones competidor</div></div>
      <div class="stat"><div class="val">{datos['total_vendidos']:,}</div><div class="lbl">Unidades vendidas</div></div>
      <div class="stat"><div class="val">${datos['precio_prom']:,}</div><div class="lbl">Precio promedio</div></div>
      <div class="stat"><div class="val">{datos['total_stock']:,}</div><div class="lbl">Stock total</div></div>
    </div>

    <div class="card">
      <h2>Análisis de Mercado</h2>
      {_md_a_html(analisis)}
    </div>

    <div class="card">
      <h2>Ideas para Instagram</h2>
      {_md_a_html(ideas)}
    </div>

    <div class="card">
      <h2>Plan de Acción — Esta Semana</h2>
      {_md_a_html(plan)}
    </div>

    </div>
    <footer>AgenciaML — agente de contenido</footer>
    </body>
    </html>
    """)

    ruta.write_text(html, encoding="utf-8")
    return ruta


# ─── Punto de entrada ─────────────────────────────────────────────────────────

def ejecutar(cliente: dict) -> bool:
    """Función llamada por el orquestador."""
    nombre    = cliente["cliente"]
    carpeta_c = cliente["_carpeta"].name.upper()
    carpeta_r = Path("reportes_ml") / carpeta_c

    vendedores = (
        cliente.get("plataformas", {})
               .get("mercadolibre", {})
               .get("vendedores_a_espiar", [])
    )
    nickname = vendedores[0] if vendedores else carpeta_c.lower()

    print(f"      Leyendo último reporte de {nickname}…")
    datos = _leer_ultimo_reporte(carpeta_r, nickname)
    if not datos:
        print(f"      [!] No se encontró reporte previo para {nickname}.")
        return False

    fecha = datetime.now()
    print("      Generando análisis con IA…")
    analisis = _llamar_claude(_prompt_analisis(datos, nombre), max_tokens=1800)

    print("      Generando ideas para Instagram…")
    ideas = _llamar_claude(_prompt_ideas_instagram(datos, nombre), max_tokens=1800)

    print("      Generando plan de acción…")
    plan = _llamar_claude(_prompt_plan_accion(datos, nombre), max_tokens=1200)

    carpeta_r.mkdir(parents=True, exist_ok=True)
    ruta_html = _guardar_html(analisis, ideas, plan, nombre, datos, carpeta_r, fecha)
    print(f"      Contenido guardado: {ruta_html}")

    return True
